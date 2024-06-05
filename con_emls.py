import email
import imaplib
from openpyxl import load_workbook
import psycopg2
from email.header import decode_header
from datetime import datetime


# Загружаем данные для подключения к почтовому ящику и БД
IMAP_HOST = "imap.mail.ru"
IMAP_USER = "***@*****.ru"
IMAP_PASS = "********"

DB_HOST = "******.************.ru"
DB_PORT = "****"
DB_USER = "*****"
DB_PASSWORD = "*******"
DB_NAME = "**_****"


def read_messages():
    mail = imaplib.IMAP4_SSL(IMAP_HOST)
    mail.login(IMAP_USER, IMAP_PASS)

    mail.select("inbox")
    _, message_numbers_raw = mail.search(None, 'FROM', '"t****@****24.ru"')
    message_numbers = message_numbers_raw[0].split()

    attachments = []

    for message_number in message_numbers:
        _, msg_data = mail.fetch(message_number, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])

                for part in msg.walk():
                    if part.get_content_maintype() == "multipart":
                        continue

                    encoded_filename = part.get_filename()
                    if encoded_filename:
                        decoded_filename_parts = decode_header(encoded_filename)
                        filename = " ".join([part.decode(enc or "utf-8") for part, enc in decoded_filename_parts])
                    else:
                        filename = None

                    payload = None
                    if filename and "xlsx" in filename:
                        payload = part.get_payload(decode=True)

                    if payload:
                        attachments.append((filename, payload))

                # Помечаем сообщение для удаления
        mail.store(message_number, '+FLAGS', '\\Deleted')

        # Удаляем помеченные сообщения
    mail.expunge()

    return attachments


def parse_and_write_to_db(attachment, conn):
    filename, content = attachment
    with open(filename, "wb") as f:
        f.write(content)
        print(filename)

    wb = load_workbook(filename)

    def string_to_hex(s):
        return ''.join([format(b, '02x') for b in s.encode('utf-8')])

    query = []

    # Цикл для чтения информации с каждого листа
    for sheet_name in wb.sheetnames:
        # Выбираем рабочий лист
        worksheet = wb[sheet_name]

        # Достаем Прибор учёта из ячеек H6:AU6
        meter = ''
        for cell in worksheet['H6':'AU6'][0]:
            meter += (cell.value if cell.value is not None else '') + ' '

        # Конвертируем meter в hex
        meter_hex = string_to_hex(meter.strip())

        # Достаем Дату из ячеек A11:A34
        dates = []
        for row in worksheet['A11':'A34']:
            for cell in row:
                dates.append(cell.value)

        # Достаем Q из ячеек AR11:AR34
        q_values = []
        for row in worksheet['AR11':'AR34']:
            for cell in row:
                q_values.append(cell.value)

        date_format = '%d.%m.%Y %H:%M'
        output_format = '%Y-%m-%d %H:%M:%S'

        for date, q in zip(dates, q_values):
            dt_convert = datetime.strptime(date, date_format)
            date = dt_convert.strftime(output_format)
            if q != float:
                try:
                    q = float(q)
                except:
                    q = float(0)
            query.append({
                'at': date,
                'device': meter_hex,
                'chart': 'Q',
                'value': q})
        # # Печатаем результат
        for record in query:
            print(record)

    conn.autocommit = True
    cur = conn.cursor()
    val = ",".join(cur.mogrify("(%s,%s,%s,%s)", (x['at'], x['device'],
                                                 x['chart'], x['value'])).decode('utf-8') for x in query)
    cur.execute("""INSERT INTO sch_roks.tbl_journal_counters VALUES """ + val)
    cur.close()
    conn.close()


def main():
    conn = psycopg2.connect(host=DB_HOST, port=DB_PORT, dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    attachments = read_messages()
    for attachment in attachments:
        parse_and_write_to_db(attachment, conn)


if __name__ == "__main__":
    main()
